# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook 
from openpyxl import Workbook
import numpy as np


#读取数据
shuju = load_workbook("2019年交通信息融合与挖掘-期末大作业数据.xlsx")
#取得sheet4
data4 = shuju[shuju.sheetnames[3]]



#1
#加权计算整条干线的平均值
speed_g = []
for i in range(3,171):
    pingjun = data4.cell(i,3).value*0.17082+data4.cell(i,4).value*0.23147\
              +data4.cell(i,5).value*0.16509+data4.cell(i,6).value*0.20216\
              +data4.cell(i,7).value*0.08996+data4.cell(i,8).value*0.1405
    speed_g.append(pingjun)

#(1)对得到的数据进行统计分析
tongji =[]
#平均值
junzhi = np.mean(speed_g)
tongji.append(junzhi)
#中列数
zhonglie = (max(speed_g)+min(speed_g))/2
tongji.append(zhonglie)
#中位数
zhongwei = np.median(speed_g)
tongji.append(zhongwei)
#标准差
wbiaozhun = np.std(speed_g,ddof=1)
tongji.append(wbiaozhun)
#变异系数
CV = wbiaozhun/junzhi
tongji.append(CV)
#最大值
maxs = max(speed_g)
tongji.append(maxs)
#最小值
mins = min(speed_g)
tongji.append(mins)
#样本数
shul = len(speed_g)
tongji.append(shul)

#(2)五分位数
wufen = []
wufen.append(mins)
wufen.append(np.percentile(speed_g,25))
wufen.append(zhongwei)
wufen.append(np.percentile(speed_g,75))
wufen.append(maxs)
#画箱图
import matplotlib.pyplot as plt
plt.boxplot(speed_g,patch_artist=True,showbox=True)
plt.show()
#(3)分布图
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus'] = False
fig = plt.figure()
ax = fig.add_subplot(111)
ax.hist(speed_g,bins=[33,35,37,39,41,43,45,47],color='steelblue',edgecolor='black')
ax.set_title('干线平均速度分布图')
ax.set_xlabel('速度km/h')
ax.set_ylabel('频数')
x=[34,36,38,40,42,44,46]
y=[5/168,26/168,66/168,114/168,151/168,166/168,168/168]
#双y轴
ax2 = ax.twinx()
ax2.plot(x,y,'.-',color='darkorange')
ax2.set_ylabel('累计频率曲线100%')
plt.show()

#打开一个新的excel表格用于写入计算结果
wr1 = Workbook()
#获取sheet1并命名
ws = wr1.active
ws.title = '第1题'
#为其赋值
#前两列时间
for row in range(1,171):
    for col in range(1,3):
        ws.cell(row,col).value=data4.cell(row,col).value
#干线速度       
ws.cell(1,3).value='平均速度km/h'
ws.cell(2,3).value='整条干线'
for row in range(3,171):
     ws.cell(row,3).value=speed_g[row-3]
#统计分析量
ws.cell(1,4).value = '统计分析'
ws.cell(2,4).value = '算术平均值'
ws.cell(3,4).value = '中列数'
ws.cell(4,4).value = '中位数'
ws.cell(5,4).value = '标准差'
ws.cell(6,4).value = '变异系数'
ws.cell(7,4).value = '最大值'
ws.cell(8,4).value = '最小值'
ws.cell(9,4).value = '样本数'
for row in range(2,10):
    ws.cell(row,5).value=tongji[row-2]
#五分位数
ws.cell(1,6).value='五分位数'
ws.cell(2,6).value='MIN'
ws.cell(3,6).value='Q1'
ws.cell(4,6).value='Median'
ws.cell(5,6).value='Q3'
ws.cell(6,6).value='MAX'
for i in range(2,7):
    ws.cell(i,7).value=wufen[i-2]
#保存数据集
wr1.save('数据集1.xlsx')



#2
data1 = shuju[shuju.sheetnames[0]]
#(1)剔除异常数据-阈值法
#五分位数原则
sudu=[]
liuliang=[]
zhanyoulv=[]
for i in range(3,171):
    for j in range(3,9):
        if data1.cell(i,j).value is not None:
            sudu.append(data1.cell(i,j).value)
    for j in range(9,15):
        if data1.cell(i,j).value is not None:
            liuliang.append(data1.cell(i,j).value)
    for j in range(15,21):
        if data1.cell(i,j).value is not None:
            zhanyoulv.append(data1.cell(i,j).value)
#计算iqr值
iqr1 = np.percentile(sudu,75)-np.percentile(sudu,25)
iqr2 = np.percentile(liuliang,75)-np.percentile(liuliang,25)
iqr3 = np.percentile(zhanyoulv,75)-np.percentile(zhanyoulv,25)
#计算正常区间
sudumin = max(min(sudu),np.percentile(sudu,25)-iqr1)
sudumax = min(max(sudu),np.percentile(sudu,75)+iqr1)
liulmin = max(min(liuliang),np.percentile(liuliang,25)-iqr2)
liulmax = min(max(liuliang),np.percentile(liuliang,75)+iqr2)
zhanylmin = max(min(zhanyoulv),np.percentile(zhanyoulv,25)-iqr3)
zhanylmax = min(max(zhanyoulv),np.percentile(zhanyoulv,75)+iqr3)
#通过五分位数法进行异常数据剔除
for i in range(3,171):
    for j in range(3,9):
        if data1.cell(i,j).value is not None:
            if data1.cell(i,j).value<sudumin or data1.cell(i,j).value>sudumax:
                data1.cell(i,j).value=None
    for j in range(9,15):
        if data1.cell(i,j).value is not None:
            if data1.cell(i,j).value<liulmin or data1.cell(i,j).value>liulmax:
                data1.cell(i,j).value=None
    for j in range(15,21):
        if data1.cell(i,j).value is not None:
            if data1.cell(i,j).value<zhanylmin or data1.cell(i,j).value>zhanylmax:
                data1.cell(i,j).value=None
#（2）补全缺失数据-时间序列数据修补
#如果缺失数据为前5个周期的数据采用线性内插法
#如果为第6个以后的周期的数据采用加权移动平均法
for i in range(3,171):
    for j in range(3,21):
        if data1.cell(i,j).value is None:
            if i==3:
                data1.cell(i,j).value=(data1.cell(i+1,j).value+data1.cell(i+2,j).value)/2
            #线性内插法
            if 3<i<8:
                #流量存为整数
                if 8<j & j<15:
                    data1.cell(i,j).value=int(data1.cell(i-1,j).value+data1.cell(i+1,j).value)
                #占有率和速度存为浮点型
                else:
                    data1.cell(i,j).value=(data1.cell(i-1,j).value+data1.cell(i+1,j).value)
            #加权移动平均法
            if i>7:
                if 8<j & j<15:
                    data1.cell(i,j).value=int((1*data1.cell(i-5,j).value+2*data1.cell(i-4,j).value\
                           +3*data1.cell(i-3,j).value+4*data1.cell(i-2,j).value\
                           +5*data1.cell(i-1,j).value)/(5+4+3+2+1))
                else:
                    data1.cell(i,j).value=(1*data1.cell(i-5,j).value+2*data1.cell(i-4,j).value\
                           +3*data1.cell(i-3,j).value+4*data1.cell(i-2,j).value\
                           +5*data1.cell(i-1,j).value)/(5+4+3+2+1)
#保存处理后的数据集
shuju.save('数据集2.xlsx')



#3
#对路段4的线圈数据聚类分析
#创建路段4的速度，流量和占有率的三维数组
shuzu4 = np.zeros([168,3])
for i in range(3,171):
    hang=[]
    hang.append(data1.cell(i,6).value)
    hang.append(data1.cell(i,12).value)
    hang.append(data1.cell(i,18).value)
    shuzu4[i-3][0]=hang[0]
    shuzu4[i-3][1]=hang[1]
    shuzu4[i-3][2]=hang[2]
#k-means聚类
from sklearn.cluster import KMeans
from sklearn.metrics import silhouette_score
#根据肘方法确定聚类簇数
#距离平方和与轮廓系数
Sum_of_squared_distances = []
si_score =[]
#轮廓系数在簇数大于1才有意义
K = range(2,15)
for k in K:
    km = KMeans(n_clusters=k)
    km = km.fit(shuzu4)
    Sum_of_squared_distances.append(km.inertia_)
    score = silhouette_score(shuzu4,km.labels_)
    si_score.append(score)

#画出结果图形
fig = plt.figure()
ax = fig.add_subplot(111)
ax.plot(K, Sum_of_squared_distances, 'x-r',label='SSE')
ax.set_title('手肘法')
ax.set_xlabel('簇数')
ax.set_ylabel('Sum_of_squared_distances')
ax2 = ax.twinx()
ax2.plot(K,si_score,'D-c',label='轮廓系数')
ax2.set_ylabel('轮廓系数')
plt.show()

#根据两者选择簇数为4进行聚类
#聚类结果可视化
from mpl_toolkits.mplot3d import Axes3D
fig = plt.figure()
ax = Axes3D(fig, rect=[0, 0, .95, 1], elev=48, azim=134)
estimators = KMeans(n_clusters=4)
estimators.fit(shuzu4)
ax.scatter(shuzu4[:,1],shuzu4[:,0],shuzu4[:,2],
           c=estimators.labels_.astype(np.float),edgecolor='k')

ax.set_xlabel('流量')
ax.set_ylabel('速度km/h')
ax.set_zlabel('时间占有率%')
ax.set_title('4簇聚类效果图')

#计算各类簇数据的统计特征
#将各簇的数据放入各自的数组
cu1=[]
cu2=[]
cu3=[]
cu4=[]
for i in range(0,168):
    if estimators.labels_[i]==0:
        cu1.append(shuzu4[i])
    elif estimators.labels_[i]==1:
        cu2.append(shuzu4[i])
    elif estimators.labels_[i]==2:
        cu3.append(shuzu4[i])
    elif estimators.labels_[i]==3:
        cu4.append(shuzu4[i])

#将嵌套列表转换为numpy数组，支持多维切片
cu1 = np.array(cu1)
cu2 = np.array(cu2)
cu3 = np.array(cu3)
cu4 = np.array(cu4)
#计算统计特征
tjtz1 = np.zeros([5,3])
for i in range(0,3):
    tjtz1[0][i]=np.mean(cu1[:,i])
    tjtz1[1][i]=np.var(cu1[:,i],ddof=1)
    tjtz1[2][i]=max(cu1[:,i])
    tjtz1[3][i]=min(cu1[:,i])
    tjtz1[4][i]=len(cu1[:,i])
tjtz2 = np.zeros([5,3])
for i in range(0,3):
    tjtz2[0][i]=np.mean(cu2[:,i])
    tjtz2[1][i]=np.var(cu2[:,i],ddof=1)
    tjtz2[2][i]=max(cu2[:,i])
    tjtz2[3][i]=min(cu2[:,i])
    tjtz2[4][i]=len(cu2[:,i])
tjtz3 = np.zeros([5,3])
for i in range(0,3):
    tjtz3[0][i]=np.mean(cu3[:,i])
    tjtz3[1][i]=np.var(cu3[:,i],ddof=1)
    tjtz3[2][i]=max(cu3[:,i])
    tjtz3[3][i]=min(cu3[:,i])
    tjtz3[4][i]=len(cu3[:,i])
tjtz4 = np.zeros([5,3])
for i in range(0,3):
    tjtz4[0][i]=np.mean(cu4[:,i])
    tjtz4[1][i]=np.var(cu4[:,i],ddof=1)
    tjtz4[2][i]=max(cu4[:,i])
    tjtz4[3][i]=min(cu4[:,i])
    tjtz4[4][i]=len(cu4[:,i])
    

#4
#相异性
#导入sheet2浮动车数据
data2 = shuju[shuju.sheetnames[1]]
#线圈路段3和6欧几里得距离  
pf13 = 0
pf16 = 0
for i in range(3,171):
    pf13 = pf13 +pow((data1.cell(i,5).value-data4.cell(i,5).value),2)
    pf16 = pf16 +pow((data1.cell(i,8).value-data4.cell(i,8).value),2)
xq3_ojld = pow(pf13,0.5)
xq6_ojld = pow(pf16,0.5)
#浮动车路段3和6的欧几里得距离
pf13 = 0
pf16 = 0
for i in range(3,171):
    pf13 = pf13 +pow((data2.cell(i,5).value-data4.cell(i,5).value),2)
    pf16 = pf16 +pow((data2.cell(i,8).value-data4.cell(i,8).value),2)
fdc3_ojld = pow(pf13,0.5)
fdc6_ojld = pow(pf16,0.5)
#路段3和路段6的dtw距离
def dtw_distance(ts_a,ts_b,d=lambda x,y:abs(x-y),mww=10000):
    import numpy as np
    #Create cost matrix via broadcasting with large int
    ts_a,ts_b=np.array(ts_a),np.array(ts_b)
    M,N=len(ts_a),len(ts_b)
    cost=np.ones((M,N))
    
    #Initialize the first row and column
    cost[0,0]=d(ts_a[0],ts_b[0])
    for i in range(1,M):
        cost[i,0] = cost[i-1,0] + d(ts_a[i],ts_b[0])
        
    for j in range(1,N):
        cost[0,j] = cost[0,j-1] + d(ts_a[0],ts_b[j])
        
    #Populate rest of cost matrix within window
    for i in range(1,M):
        for j in range(max(1,i-mww),min(N,i+mww)):
            choices = cost[i-1,j-1],cost[i,j-1],cost[i-1,j]
            cost[i,j] = min(choices)+d(ts_a[i],ts_b[j])
            
    
    #Return DTW distance given window
    return cost[-1,-1]
#提出各检测器各路段的速度序列
xq3sudu = []
xq6sudu = []
fdc3sudu = []
fdc6sudu = []
zs3sudu = []
zs6sudu = []
for i in range(3,171):
    xq3sudu.append(data1.cell(i,5).value)
    xq6sudu.append(data1.cell(i,8).value)
    fdc3sudu.append(data2.cell(i,5).value)
    fdc6sudu.append(data2.cell(i,8).value)
    zs3sudu.append(data4.cell(i,5).value)
    zs6sudu.append(data4.cell(i,8).value)
#利用dtw函数计算dtw距离
#线圈的dtw距离
xq3_dtw = dtw_distance(xq3sudu,zs3sudu)
xq6_dtw = dtw_distance(xq6sudu,zs6sudu)
#浮动车的dtw距离
fdc3_dtw = dtw_distance(fdc3sudu,zs3sudu)
fdc6_dtw = dtw_distance(fdc6sudu,zs6sudu)

#相关性
#计算矩阵(XY),(X平方等)
xqzs3 = []
xqzs6 = []
fdczs3 = []
fdczs6 = []
xq3pf = []
xq6pf =[]
fdc3pf = []
fdc6pf = []
zs3pf = []
zs6pf = []
for i in range(0,168):
    xqzs3.append(xq3sudu[i]*zs3sudu[i])
    xqzs6.append(xq6sudu[i]*zs6sudu[i])
    fdczs3.append(fdc3sudu[i]*zs3sudu[i])
    fdczs6.append(fdc6sudu[i]*zs6sudu[i])
    xq3pf.append(pow(xq3sudu[i],2))
    xq6pf.append(pow(xq6sudu[i],2))
    fdc3pf.append(pow(fdc3sudu[i],2))
    fdc6pf.append(pow(fdc6sudu[i],2))
    zs3pf.append(pow(zs3sudu[i],2))
    zs6pf.append(pow(zs6sudu[i],2))
#计算期望E(X),E(Y)，E(XY)
qw =np.zeros([2,8])
qw[0][0] = np.mean(xq3sudu)
qw[0][1] = np.mean(fdc3sudu)
qw[0][2] = np.mean(zs3sudu)
qw[0][3] = np.mean(xqzs3)
qw[0][4] = np.mean(fdczs3)
qw[0][5] = np.mean(xq3pf)
qw[0][6] = np.mean(fdc3pf)
qw[0][7] = np.mean(zs3pf)
qw[1][0] = np.mean(xq6sudu)
qw[1][1] = np.mean(fdc6sudu)
qw[1][2] = np.mean(zs6sudu)
qw[1][3] = np.mean(xqzs6)
qw[1][4] = np.mean(fdczs6)
qw[1][5] = np.mean(xq6pf)
qw[1][6] = np.mean(fdc6pf)
qw[1][7] = np.mean(zs6pf)
#计算协方差
xq3_cov = qw[0][3]-qw[0][0]*qw[0][2]
xq6_cov = qw[1][3]-qw[1][0]*qw[1][2]
fdc3_cov = qw[0][4]-qw[0][1]*qw[0][2]
fdc6_cov = qw[1][4]-qw[1][1]*qw[1][2]
#计算相关系数 
y3sigma = pow(qw[0][7]-pow(qw[0][2],2),0.5)
y6sigma = pow(qw[1][7]-pow(qw[1][2],2),0.5)
xq3_xgxs = xq3_cov/(pow(qw[0][5]-pow(qw[0][0],2),0.5)*y3sigma)
xq6_xgxs = xq6_cov/(pow(qw[1][5]-pow(qw[1][0],2),0.5)*y6sigma)
fdc3_xgxs = fdc3_cov/(pow(qw[0][6]-pow(qw[0][1],2),0.5)*y3sigma)
fdc6_xgxs = fdc6_cov/(pow(qw[0][6]-pow(qw[0][1],2),0.5)*y6sigma)



#5
data3 = shuju[shuju.sheetnames[2]]
data2 = shuju[shuju.sheetnames[1]]
data1 = shuju[shuju.sheetnames[0]]
#定义RMSE函数
def rmse(a,b):
    a,b = np.array(a),np.array(b)
    pingfang = 0
    for i in range(0,len(a)):
        pingfang = pingfang+pow(a[i]-b[i],2)
    zhi = pow(pingfang/len(a),0.5)
    return zhi
#定义MAPE函数
def mape(a,b):
    import numpy as np
    a,b = np.array(a),np.array(b)
    juedui = 0
    for i in range(0,len(a)):
        juedui = juedui+abs((b[i]-a[i])/b[i])
    zhi = juedui/len(a)
    return zhi
#定义一个将sheet表中一列数据读出的函数
def duchu(datax,lie,hangf,hange):
    dedao = []
    for i in range(hangf,hange+1):
            dedao.append(datax.cell(i,lie).value)
    return dedao
#将需要的速度读成列表保存
xqsudu =[duchu(data1,3,3,170),duchu(data1,4,3,170),duchu(data1,5,3,170),duchu(data1,6,3,170),\
         duchu(data1,7,3,170),duchu(data1,8,3,170)]
fdcsudu =[duchu(data2,3,3,170),duchu(data2,4,3,170),duchu(data2,5,3,170),duchu(data2,6,3,170),\
         duchu(data2,7,3,170),duchu(data2,8,3,170)]
avisudu = duchu(data3,3,3,170)
zssudu =[duchu(data4,3,3,170),duchu(data4,4,3,170),duchu(data4,5,3,170),duchu(data4,6,3,170),\
         duchu(data4,7,3,170),duchu(data4,8,3,170)]
#线圈的估计误差
#RMSE
total_ = 0
xq_rmse_ = []
for i in range(0,6):
    total_ = total_ + rmse(xqsudu[i],zssudu[i])
    xq_rmse_.append(rmse(xqsudu[i],zssudu[i]))
xq_rmse = total_/6
#mape
xq_mape_=[]
total_ = 0
for i in range(0,6):
    total_ = total_ + mape(xqsudu[i],zssudu[i])
    xq_mape_.append(mape(xqsudu[i],zssudu[i]))
xq_mape = total_/6
#浮动车的估计误差
#RMSE
total_ = 0
fdc_rmse_ = []
for i in range(0,6):
    total_ = total_ + rmse(fdcsudu[i],zssudu[i])
    fdc_rmse_.append(rmse(fdcsudu[i],zssudu[i]))
fdc_rmse = total_/6
#MAPE
fdc_mape_ = []
total_ = 0
for i in range(0,6):
    total_ = total_ + mape(fdcsudu[i],zssudu[i])
    fdc_mape_.append(mape(fdcsudu[i],zssudu[i]))
fdc_mape = total_/6
#AVI的估计误差
#RMSE
avi_rmse = rmse(avisudu,speed_g)
#MAPE
avi_mape = mape(avisudu,speed_g)



#6信息融合
#将AVI数据根据线圈数据进行划分，划分为6个路段
#每路段长度和总长度
length_l = [0.507,0.687,0.490,0.600,0.267,0.417,2.968]
#建立avi的空集
avi = np.zeros([168,6])
for i in range(3,171):
    #每行的每个路段的时间及总时间
    t_ = [0,0,0,0,0,0]
    t_sum= 0
    for j in range(0,6):
        t_[j]=length_l[j]/data1.cell(i,j+3).value
        t_sum = t_sum+t_[j]
    #按照线圈数据的时间分配avi的时间
    #avi速度得到的路段行程时间，计算出各路段的估计速度
    t_avi = length_l[6]/data3.cell(i,3).value
    for j in range(0,6):
        avi[i-3,j]=t_avi*(t_[j]/t_sum)
        avi[i-3,j]=length_l[j]/avi[i-3,j]



#根据加权平均法
#权重系数的计算
quanzhong1 = []
for i in range(0,6):
    quanzhong1.append(1-mape(xqsudu[i][0:120],zssudu[i][0:120]))
quanzhong2 = []
for i in range(0,6):
    quanzhong2.append(1-mape(fdcsudu[i][0:120],zssudu[i][0:120]))
quanzhong3 = []
for i in range(0,6):
    quanzhong3.append(1-mape(avi[0:120,i],zssudu[i][0:120]))

#后4个小时的预测值函数
#定义预测函数，预测方法1是直接三种数据融合
def yuce1(luduan,xq,fdc,avi):
    zong = quanzhong1[luduan-1]+quanzhong2[luduan-1]+quanzhong3[luduan-1]
    zhi = xq*(quanzhong1[luduan-1]/zong)+fdc*(quanzhong2[luduan-1]/zong)\
         +avi*(quanzhong3[luduan-1]/zong)
    return zhi

#预测方法2是将每个路段的权重进行筛选，选择出最高的两个，将其数据进行融合
def yuce2(luduan,xq,fdc,avi):
    quanzhong = [quanzhong1[luduan-1],quanzhong2[luduan-1],quanzhong3[luduan-1]]
    paoqi = quanzhong.index(min(quanzhong))
    if paoqi== 0:
         he = quanzhong2[luduan-1]+quanzhong3[luduan-1]
         zhi=fdc*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif paoqi==1:
         he = quanzhong1[luduan-1]+quanzhong3[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif paoqi==2:
         he = quanzhong1[luduan-1]+quanzhong2[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+fdc*(quanzhong3[luduan-1]/he)
    return zhi
#预测方法3是当其准确度都不小于0.8时，采用该检测数据
#但如果最小值小于0.8时，就选取较高的两种检测数据融合
def yuce3(luduan,xq,fdc,avi):
    quanzhong = [quanzhong1[luduan-1],quanzhong2[luduan-1],quanzhong3[luduan-1]]
    paoqi = quanzhong.index(min(quanzhong))
    if min(quanzhong) >= 0.8:
         he = quanzhong[0]+quanzhong[1]+quanzhong[2]
         zhi=xq*(quanzhong1[luduan-1]/he)+fdc*(quanzhong2[luduan-1]/he)\
         +avi*(quanzhong3[luduan-1]/he)
    elif min(quanzhong) < 0.8  and  paoqi== 0:
         he = quanzhong2[luduan-1]+quanzhong3[luduan-1]
         zhi=fdc*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif min(quanzhong) < 0.8  and  paoqi==1:
         he = quanzhong1[luduan-1]+quanzhong3[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif min(quanzhong) < 0.8  and  paoqi==2:
         he = quanzhong1[luduan-1]+quanzhong2[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+fdc*(quanzhong3[luduan-1]/he)
    return zhi

#用法1和法2得到后4个小时的融合值,并计算mape
ronghe1=[]
wucha1=[]
ronghe2=[]
wucha2=[]
for j in range(0,6):
    yucezhi1=[]
    yucezhi2=[]
    for i in range(120,168):
        dedao1 = yuce1(j+1,xqsudu[j][i],fdcsudu[j][i],avi[i,j])
        yucezhi1.append(dedao1)
        dedao2 = yuce2(j+1,xqsudu[j][i],fdcsudu[j][i],avi[i,j])
        yucezhi2.append(dedao2)
    ronghe1.append(yucezhi1)
    ronghe2.append(yucezhi2)
    wuchal = mape(yucezhi1,zssudu[j][120:168])
    wucha2l = mape(yucezhi2,zssudu[j][120:168])
    wucha1.append(wuchal)
    wucha2.append(wucha2l)

#利用方法3计算融合值
wucha = []
ronghe = []
for j in range(0,6):
    yucezhi=[]
    for i in range(120,168):
        dedao = yuce3(j+1,xqsudu[j][i],fdcsudu[j][i],avi[i,j])
        yucezhi.append(dedao)
    ronghe.append(yucezhi)
    wuchall = mape(yucezhi,zssudu[j][120:168])
    wucha.append(wuchall)

#打开一个新的excel表格用于写入计算结果
wr6 = Workbook()
#获取sheet1并命名
ws6 = wr6.active
ws6.title = '第6题预测结果'
#为其赋值
#前两列时间
for row in range(1,3):
    for col in range(1,3):
        ws6.cell(row,col).value=data4.cell(row,col).value
for row in range(3,51):
    for col in range(1,3):
        ws6.cell(row,col).value=data4.cell(row+120,col).value
ws6.merge_cells('C1:H1')
ws6.cell(1,3).value = '平均速度km/h'
ws6.cell(2,3).value = '路段1'
ws6.cell(2,4).value = '路段2'
ws6.cell(2,5).value = '路段3'
ws6.cell(2,6).value = '路段4'
ws6.cell(2,7).value = '路段5'
ws6.cell(2,8).value = '路段6'
for row in range(3,51):
    for col in range(3,9):
        ws6.cell(row,col).value = ronghe[col-3][row-3]
wr6.save('数据集6.xlsx')