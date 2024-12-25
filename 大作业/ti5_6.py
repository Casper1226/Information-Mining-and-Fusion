# -*- coding: utf-8 -*-
from openpyxl.reader.excel import load_workbook 
import numpy as np


#读取数据
shuju = load_workbook("2019年交通信息融合与挖掘-期末大作业数据.xlsx")
#取得sheet4
data4 = shuju[shuju.sheetnames[3]]
#加权计算整条干线的平均值
speed_g = []
for i in range(3,171):
    pingjun = data4.cell(i,3).value*0.17082+data4.cell(i,4).value*0.23147\
              +data4.cell(i,5).value*0.16509+data4.cell(i,6).value*0.20216\
              +data4.cell(i,7).value*0.08996+data4.cell(i,8).value*0.1405
    speed_g.append(pingjun)


#5
data3 = shuju[shuju.sheetnames[2]]
data2 = shuju[shuju.sheetnames[1]]
data1 = shuju[shuju.sheetnames[0]]
#定义RMSE函数
def rmse(a,b):
    a,b = np.array(a),np.array(b)
    pingfang = 0
    for i in range(0,len(a)):
        pingfang = pingfang+pow(a[i]-b[i],2)
    zhi = pow(pingfang/len(a),0.5)
    return zhi
#定义MAPE函数
def mape(a,b):
    import numpy as np
    a,b = np.array(a),np.array(b)
    juedui = 0
    for i in range(0,len(a)):
        juedui = juedui+abs((b[i]-a[i])/b[i])
    zhi = juedui/len(a)
    return zhi
#定义一个将sheet表中一列数据读出的函数
def duchu(datax,lie,hangf,hange):
    dedao = []
    for i in range(hangf,hange+1):
            dedao.append(datax.cell(i,lie).value)
    return dedao
#将需要的速度读成列表保存
xqsudu =[duchu(data1,3,3,170),duchu(data1,4,3,170),duchu(data1,5,3,170),duchu(data1,6,3,170),\
         duchu(data1,7,3,170),duchu(data1,8,3,170)]
fdcsudu =[duchu(data2,3,3,170),duchu(data2,4,3,170),duchu(data2,5,3,170),duchu(data2,6,3,170),\
         duchu(data2,7,3,170),duchu(data2,8,3,170)]
avisudu = duchu(data3,3,3,170)
zssudu =[duchu(data4,3,3,170),duchu(data4,4,3,170),duchu(data4,5,3,170),duchu(data4,6,3,170),\
         duchu(data4,7,3,170),duchu(data4,8,3,170)]
#线圈的估计误差
#RMSE
total_ = 0
xq_rmse_ = []
for i in range(0,6):
    total_ = total_ + rmse(xqsudu[i],zssudu[i])
    xq_rmse_.append(rmse(xqsudu[i],zssudu[i]))
xq_rmse = total_/6
#mape
xq_mape_=[]
total_ = 0
for i in range(0,6):
    total_ = total_ + mape(xqsudu[i],zssudu[i])
    xq_mape_.append(mape(xqsudu[i],zssudu[i]))
xq_mape = total_/6
#浮动车的估计误差
#RMSE
total_ = 0
fdc_rmse_ = []
for i in range(0,6):
    total_ = total_ + rmse(fdcsudu[i],zssudu[i])
    fdc_rmse_.append(rmse(fdcsudu[i],zssudu[i]))
fdc_rmse = total_/6
#MAPE
fdc_mape_ = []
total_ = 0
for i in range(0,6):
    total_ = total_ + mape(fdcsudu[i],zssudu[i])
    fdc_mape_.append(mape(fdcsudu[i],zssudu[i]))
fdc_mape = total_/6
#AVI的估计误差
#RMSE
avi_rmse = rmse(avisudu,speed_g)
#MAPE
avi_mape = mape(avisudu,speed_g)


#6信息融合
#将AVI数据根据线圈数据进行划分，划分为6个路段
#每路段长度和总长度
length_l = [0.507,0.687,0.490,0.600,0.267,0.417,2.968]
#建立avi的空集
avi = np.zeros([178,6])
for i in range(3,171):
    #每行的每个路段的时间及总时间
    t_ = [0,0,0,0,0,0]
    t_sum= 0
    for j in range(0,6):
        t_[j]=length_l[j]/data1.cell(i,j+3).value
        t_sum = t_sum+t_[j]
    #按照线圈数据的时间分配avi的时间
    #avi速度得到的路段行程时间
    t_avi = length_l[6]/data3.cell(i,3).value
    for j in range(0,6):
        avi[i-3,j]=t_avi*(t_[j]/t_sum)
        avi[i-3,j]=length_l[j]/avi[i-3,j]


total_ = 0
for i in range(0,6):
    total_ = total_ + mape(avi[0:168,i],zssudu[i])
jisuan_mape = total_/6

#根据加权平均法
#权重系数的选择
quanzhong1 = []
for i in range(0,6):
    quanzhong1.append(1-mape(xqsudu[i][0:120],zssudu[i][0:120]))
quanzhong2 = []
for i in range(0,6):
    quanzhong2.append(1-mape(fdcsudu[i][0:120],zssudu[i][0:120]))
quanzhong3 = []
for i in range(0,6):
    quanzhong3.append(1-mape(avi[0:120,i],zssudu[i][0:120]))

#后4个小时的预测值函数
def yuce(luduan,xq,fdc,avi):
    quanzhong = [quanzhong1[luduan-1],quanzhong2[luduan-1],quanzhong3[luduan-1]]
    paoqi = quanzhong.index(min(quanzhong))
    if paoqi== 0:
         he = quanzhong2[luduan-1]+quanzhong3[luduan-1]
         zhi=fdc*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif paoqi==1:
         he = quanzhong1[luduan-1]+quanzhong3[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+avi*(quanzhong3[luduan-1]/he)
    elif paoqi==2:
         he = quanzhong1[luduan-1]+quanzhong2[luduan-1]
         zhi=xq*(quanzhong2[luduan-1]/he)+fdc*(quanzhong3[luduan-1]/he)
    return zhi

#得到后4个小时的融合值,并计算mape
ronghe=[]
wucha=[]
for j in range(0,6):
    yucezhi=[]
    for i in range(120,168):
        dedao = yuce(j+1,xqsudu[j][i],fdcsudu[j][i],avi[i,j])
        yucezhi.append(dedao)
    ronghe.append(yucezhi)
    wuchal = mape(yucezhi,zssudu[j][120:168])
    wucha.append(wuchal) 
#得到平均mape
pingjun = np.mean(wucha)

#未融合时线圈后4个小时的值的平均误差
xq4mape =[]
for i in range(0,6):
    xq4mape.append(mape(xqsudu[i][123:168],zssudu[i][123:168]))
